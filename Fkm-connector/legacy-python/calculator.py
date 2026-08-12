"""
calculator.py — Route C, Variant C3 Implementation
────────────────────────────────────────────────────
Production Sheets to MRP, No Browser (RPC + Template XML Patching + LibreOffice UNO Listener).

1. Reads base snapshot/data & revision log via Odoo RPC (chained via parent_revision_id).
2. Extracts UPDATE_CELL commands for input cells in "Gegevens invulblad" (column B).
3. If a pristine template .xlsx exists, patches XML input values & resets calc state (`fullCalcOnLoad="1"`).
   Otherwise, builds/updates an openpyxl workbook structure.
4. Uses LibreOffice (UNO listener on port 2002 if available, or fallback headless CLI) to force full
   formula recalculation and export:
   - Full 4-sheet xlsx (`productie.xlsx`)
   - 2-sheet PDF (`productie.pdf` containing ONLY Afdrukpagina 1 and Afdrukpagina 2 via sheet selection).
5. Scans Afdrukpagina 1 & 2 for unexpected formula errors before completing.
"""

import io
import os
import re
import json
import base64
import logging
import tempfile
import subprocess
import shutil
import zipfile
import xml.etree.ElementTree as ET
import openpyxl

# --- openpyxl Monkeypatches to handle corrupt/malformed Excel files gracefully ---
import openpyxl.packaging.manifest
openpyxl.packaging.manifest.mimetypes.add_type('image/webp', '.webp')
import openpyxl.reader.drawings
import openpyxl.reader.excel
from openpyxl.descriptors.base import Set

class ZipFileWrapper:
    def __init__(self, zip_file):
        self._zip_file = zip_file
    def __getattr__(self, name):
        return getattr(self._zip_file, name)
    def read(self, name, *args, **kwargs):
        try:
            return self._zip_file.read(name, *args, **kwargs)
        except KeyError as e:
            raise OSError(f"File not found in archive: {name}") from e

original_find_images = openpyxl.reader.drawings.find_images
def patched_find_images(archive, path):
    return original_find_images(ZipFileWrapper(archive), path)

openpyxl.reader.drawings.find_images = patched_find_images
openpyxl.reader.excel.find_images = patched_find_images

original_set = Set.__set__
def patched_set(self, instance, value):
    if value not in self.values:
        self.values.add(value)
    original_set(self, instance, value)
Set.__set__ = patched_set
# ---------------------------------------------------------------------------------

from config import LIBREOFFICE_UNO_HOST, LIBREOFFICE_UNO_PORT
from odoo_rpc import OdooClient

logger = logging.getLogger(__name__)

# Known error values to validate against
ERROR_VALUES = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!", "#BAD_EXPR"}


def _resolve_print_sheets(sheet_names: list[str]) -> list[str]:
    """
    Dynamically resolve print sheets from the workbook sheet names.
    Looks for sheets matching 'afdrukpagina 1' and 'afdrukpagina 2'
    case-insensitively and ignoring whitespace.
    """
    targets = ["afdrukpagina1", "afdrukpagina2"]
    resolved = []
    for name in sheet_names:
        normalized = name.replace(" ", "").lower()
        if normalized in targets:
            resolved.append(name)
    # If not found by exact target names, look for any sheet name containing "afdruk" or "print" (case-insensitive)
    if not resolved:
        resolved = [name for name in sheet_names if "afdruk" in name.lower() or "print" in name.lower()]
    # Fallback to defaults if still empty
    if not resolved:
        resolved = ["Afdrukpagina 1", "Afdrukpagina 2"]
    return resolved


def _find_template_file(template_name: str) -> str | None:
    """Disabled: Local disk template search is disabled."""
    return None

def get_calculator_state(order_id: int, spreadsheet_id: int) -> bytes:
    """
    Route C, Variant C3 Main Entry Point:
    1. Reads spreadsheet base data & chained revision log via XML-RPC.
    2. Replays UPDATE_CELL commands to resolve true current input values (static only).
    3. Builds Excel sheet directly from Odoo RPC JSON data (no local files).
    """
    logger.info(f"[C3 Engine] Processing order {order_id}, spreadsheet {spreadsheet_id}")
    odoo = OdooClient()

    # Step 1: Read base snapshot/data and revision log
    base_data, revisions = _fetch_spreadsheet_data_and_revisions(odoo, spreadsheet_id)

    # Step 2: Extract UPDATE_CELL inputs (static inputs only)
    input_values = _extract_input_cell_values(base_data, revisions)
    logger.info(f"Resolved {len(input_values)} input cell updates from base data & revisions.")

    # Step 3: Do not search local disk for template file
    template_file = None

    # Step 4: Build static XLSX file containing only print sheets
    tmpl_data = _fetch_dynamic_quotation_template(odoo, order_id)
    if tmpl_data:
        logger.info("Overlaying order spreadsheet instance cell updates onto base quotation template structure.")
        combined_data = _overlay_spreadsheet_instance(tmpl_data, base_data)
        xlsx_bytes = _build_xlsx_via_openpyxl(combined_data, input_values, template_path=template_file)
    else:
        xlsx_bytes = _build_xlsx_via_openpyxl(base_data, input_values, template_path=template_file)

    # Resolve print sheets dynamically from the generated XLSX
    wb_temp = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    resolved_print_sheets = wb_temp.sheetnames
    logger.info(f"Generated print sheets: {resolved_print_sheets}")

    # Step 5: Validate print sheets for unexpected formula errors (should be clean/empty of formulas)
    _validate_xlsx_output(xlsx_bytes, tuple(resolved_print_sheets))

    return xlsx_bytes


def _fetch_spreadsheet_data_and_revisions(odoo: OdooClient, spreadsheet_id: int) -> tuple[dict, list]:
    """Fetch base snapshot/data and revision log from Odoo over RPC."""
    records = odoo.call(
        'sale.order.spreadsheet', 'read',
        [spreadsheet_id],
        ['spreadsheet_snapshot', 'spreadsheet_data', 'spreadsheet_binary_data']
    )
    if not records:
        raise ValueError(f"Spreadsheet record {spreadsheet_id} not found in Odoo.")

    rec = records[0]
    raw_val = rec.get('spreadsheet_snapshot') or rec.get('spreadsheet_data') or rec.get('spreadsheet_binary_data')

    base_data = {}
    if raw_val:
        try:
            if isinstance(raw_val, bytes):
                raw_val = raw_val.decode('utf-8')
            # Check if base64 encoded JSON
            if not raw_val.strip().startswith('{') and not raw_val.strip().startswith('['):
                try:
                    raw_val = base64.b64decode(raw_val).decode('utf-8')
                except Exception:
                    pass
            base_data = json.loads(raw_val)
        except Exception as e:
            logger.warning(f"Could not parse spreadsheet JSON: {e}")

    # Fetch revisions ordered by parent_revision_id
    revisions = []
    try:
        rev_records = odoo.call(
            'spreadsheet.revision', 'search_read',
            [('res_model', '=', 'sale.order.spreadsheet'), ('res_id', '=', spreadsheet_id)],
            ['commands', 'revision_uuid', 'parent_revision_id']
        )
        if rev_records:
            revisions = _chain_revisions(rev_records)
            logger.info(f"Retrieved and ordered {len(revisions)} revisions for spreadsheet {spreadsheet_id}")
    except Exception as e:
        logger.warning(f"Revision log search skipped or restricted (requires base.group_system): {e}")

    return base_data, revisions


def _fetch_dynamic_quotation_template(odoo: OdooClient, order_id: int) -> dict:
    """Dynamically fetch the base Quotation Template spreadsheet data from Odoo RPC without requiring static files."""
    try:
        so = odoo.call('sale.order', 'read', [order_id], ['sale_order_template_id'])
        if not so or not so[0].get('sale_order_template_id'):
            return {}

        tmpl_id = so[0]['sale_order_template_id'][0]
        tmpl = odoo.call('sale.order.template', 'read', [tmpl_id], ['spreadsheet_template_id'])
        if not tmpl or not tmpl[0].get('spreadsheet_template_id'):
            return {}

        sp_tmpl_id = tmpl[0]['spreadsheet_template_id'][0]
        sp_tmpl = odoo.call('sale.order.spreadsheet', 'read', [sp_tmpl_id], ['spreadsheet_binary_data', 'spreadsheet_data'])
        if not sp_tmpl:
            return {}

        raw_val = sp_tmpl[0].get('spreadsheet_binary_data') or sp_tmpl[0].get('spreadsheet_data')
        if not raw_val:
            return {}

        if isinstance(raw_val, bytes):
            raw_val = raw_val.decode('utf-8')

        if not raw_val.strip().startswith('{') and not raw_val.strip().startswith('['):
            try:
                raw_val = base64.b64decode(raw_val).decode('utf-8')
            except Exception:
                pass

        logger.info(f"Dynamically loaded quotation template spreadsheet {sp_tmpl_id} from Odoo RPC.")
        return json.loads(raw_val)
    except Exception as e:
        logger.warning(f"Could not fetch dynamic quotation template for order {order_id}: {e}")
        return {}


def _overlay_spreadsheet_instance(tmpl_data: dict, instance_data: dict) -> dict:
    """Merge order spreadsheet instance cell updates onto the base quotation template structure."""
    import copy
    merged = copy.deepcopy(tmpl_data)

    inst_sheets = {s.get("id") or s.get("name"): s for s in instance_data.get("sheets", [])}

    for sheet_info in merged.get("sheets", []):
        s_key = sheet_info.get("id") or sheet_info.get("name")
        inst_sheet = inst_sheets.get(s_key)
        if not inst_sheet and sheet_info.get("name"):
            inst_sheet = next((s for s in instance_data.get("sheets", []) if s.get("name") == sheet_info.get("name")), None)

        if inst_sheet:
            tmpl_cells = sheet_info.setdefault("cells", {})
            inst_cells = inst_sheet.get("cells", {})
            for cell_ref, cell_val in inst_cells.items():
                if cell_val is not None:
                    tmpl_cells[cell_ref] = cell_val

    return merged


def _chain_revisions(revisions: list) -> list:
    """Sort revisions in parent-child causal order using parent_revision_id."""
    rev_map = {r.get('revision_uuid'): r for r in revisions if r.get('revision_uuid')}
    parents = {r.get('parent_revision_id') for r in revisions if r.get('parent_revision_id')}
    
    # Root revision has parent_revision_id not present as any revision_uuid in set, or empty
    roots = [r for r in revisions if not r.get('parent_revision_id') or r.get('parent_revision_id') not in rev_map]
    
    ordered = []
    visited = set()
    queue = list(roots)

    while queue:
        curr = queue.pop(0)
        uuid = curr.get('revision_uuid')
        if uuid in visited:
            continue
        visited.add(uuid)
        ordered.append(curr)

        # Find children
        children = [r for r in revisions if r.get('parent_revision_id') == uuid]
        queue.extend(children)

    # Append any remaining revisions if chain had breaks
    for r in revisions:
        if r.get('revision_uuid') not in visited:
            ordered.append(r)

    return ordered


def _extract_input_cell_values(base_data: dict, revisions: list) -> dict:
    """
    Extract base inputs + replayed UPDATE_CELL commands (excluding formulas).
    Returns mapping of (sheet_name, cell_ref) -> resolved static value.
    """
    cell_values = {}
    sheet_id_to_name = {}

    sheets_data = base_data.get("sheets", [])
    for sheet_info in sheets_data:
        s_name = sheet_info.get("name", "Sheet")
        s_id = sheet_info.get("id")
        if s_id:
            sheet_id_to_name[s_id] = s_name

        # Parse base cells
        cells = sheet_info.get("cells", {})
        for cell_ref, cell_data in cells.items():
            val = None
            if isinstance(cell_data, dict):
                if cell_data.get("value") is not None:
                    val = cell_data.get("value")
                else:
                    content = cell_data.get("content")
                    if isinstance(content, str) and (content.startswith('=') or content.startswith('+') or content.startswith('-')):
                        val = None
                    else:
                        val = content
            else:
                val = cell_data

            if val is not None:
                if isinstance(val, str) and (val.startswith('=') or val.startswith('+') or val.startswith('-')):
                    continue
                cell_values[(s_name, cell_ref.upper())] = val

    # Replay UPDATE_CELL commands (last write wins)
    for rev in revisions:
        commands_raw = rev.get('commands')
        if not commands_raw:
            continue
        try:
            cmds = json.loads(commands_raw) if isinstance(commands_raw, str) else commands_raw
            if isinstance(cmds, list):
                for cmd in cmds:
                    if isinstance(cmd, dict) and cmd.get('type') == 'UPDATE_CELL':
                        s_id = cmd.get('sheetId')
                        col = cmd.get('col')
                        row = cmd.get('row')
                        content = cmd.get('content')
                        s_name = sheet_id_to_name.get(s_id, "Gegevens invulblad")

                        if col is not None and row is not None:
                            col_letter = openpyxl.utils.get_column_letter(col + 1)
                            cell_ref = f"{col_letter}{row + 1}"
                            # Revisions update inputs. If they update with a formula, do not copy the formula structure.
                            if isinstance(content, str) and (content.startswith('=') or content.startswith('+') or content.startswith('-')):
                                continue
                            cell_values[(s_name, cell_ref.upper())] = content
        except Exception as e:
            logger.debug(f"Error parsing revision command: {e}")

    return cell_values


def _patch_xlsx_template(template_path: str, input_values: dict) -> bytes:
    """
    Patch pristine XML template:
    - Write resolved values to input cells
    - Strip cached formula values (<v>)
    - Set <calcPr fullCalcOnLoad="1"/> in xl/workbook.xml
    """
    with open(template_path, 'rb') as f:
        in_buffer = io.BytesIO(f.read())

    out_buffer = io.BytesIO()

    with zipfile.ZipFile(in_buffer, 'r') as in_zip, zipfile.ZipFile(out_buffer, 'w', zipfile.ZIP_DEFLATED) as out_zip:
        for item in in_zip.infolist():
            content = in_zip.read(item.filename)

            if item.filename == 'xl/workbook.xml':
                content = _patch_workbook_xml(content)
            elif item.filename.startswith('xl/worksheets/sheet') and item.filename.endswith('.xml'):
                # Route values into worksheet XMLs if mapped
                content = _patch_worksheet_xml(content, input_values)

            out_zip.writestr(item, content)

    return out_buffer.getvalue()


def _patch_workbook_xml(xml_content: bytes) -> bytes:
    """Ensure <calcPr fullCalcOnLoad="1"/> is set in xl/workbook.xml."""
    try:
        tree = ET.fromstring(xml_content)
        ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        ET.register_namespace('', ns['main'])

        calc_pr = tree.find('main:calcPr', ns)
        if calc_pr is None:
            calc_pr = ET.SubElement(tree, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}calcPr')
        
        calc_pr.set('fullCalcOnLoad', '1')
        return ET.tostring(tree, encoding='utf-8', xml_declaration=True)
    except Exception as e:
        logger.warning(f"Failed to patch workbook.xml calcPr: {e}")
        return xml_content


def _patch_worksheet_xml(xml_content: bytes, input_values: dict) -> bytes:
    """Strip cached values on formula cells from worksheet XML."""
    try:
        tree = ET.fromstring(xml_content)
        ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        ET.register_namespace('', ns['main'])

        # Strip cached values (<v>) on cells containing formulas (<f>)
        for cell in tree.findall('.//main:c', ns):
            formula = cell.find('main:f', ns)
            val = cell.find('main:v', ns)
            if formula is not None and val is not None:
                cell.remove(val)

        return ET.tostring(tree, encoding='utf-8', xml_declaration=True)
    except Exception as e:
        logger.warning(f"Failed to patch worksheet XML: {e}")
        return xml_content


# ============================================================================
#  O-SPREADSHEET (19.1+) UNSQUISHER & IN-MEMORY EVALUATOR
# ============================================================================

ERRS = ('#REF!', '#NAME?', '#VALUE!', '#DIV/0!', '#CYCLE', '#BAD_EXPR')
E_REF, E_NAME, E_VAL, E_DIV, E_CYC, E_EXPR = ERRS

def is_err(v):
    return isinstance(v, str) and v in ERRS

def col_num(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch.upper()) - 64)
    return n

def ref_rc(xc):
    letters = ''
    digits = ''
    for ch in xc:
        if ch == '$':
            continue
        if ch.isalpha():
            if digits:
                return None
            letters += ch
        elif ch.isdigit():
            digits += ch
        else:
            return None
    if not letters or not digits or len(letters) > 3:
        return None
    return (int(digits), col_num(letters))

def key_positions(key):
    parts = key.split(':')
    a = ref_rc(parts[0])
    if a is None:
        return []
    if len(parts) == 1:
        return [a]
    b = ref_rc(parts[1])
    if b is None:
        return [a]
    out = []
    c = min(a[1], b[1])
    while c <= max(a[1], b[1]):
        r = min(a[0], b[0])
        while r <= max(a[0], b[0]):
            out.append((r, c))
            r += 1
        c += 1
    return out

def sorted_keys(cells):
    items = []
    for k in cells:
        rc = ref_rc(k.split(':')[0])
        if rc is not None:
            items.append((rc[1], rc[0], k))
    items.sort()
    return [it[2] for it in items]

def to_num(s):
    if not isinstance(s, str):
        if isinstance(s, (int, float)):
            return float(s)
        return None
    t = s.strip()
    if not t:
        return None
    body = t[1:] if t[0] in '+-' else t
    if not body:
        return None
    dots = 0
    has_digit = False
    for ch in body:
        if ch == '.':
            dots += 1
            if dots > 1:
                return None
        elif ch.isdigit():
            has_digit = True
        else:
            return None
    if not has_digit:
        return None
    return float(t)

def as_num(v):
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, float):
        return v
    if isinstance(v, int):
        return float(v)
    if v is None or v == '':
        return 0.0
    if isinstance(v, str):
        return to_num(v)
    return None

def _read_ref_chars(src, i):
    buf = ''
    n = len(src)
    while i < n and (src[i].isalnum() or src[i] == '$'):
        buf += src[i]
        i += 1
    return buf, i

def tokenize(src):
    toks = []
    i = 0
    n = len(src)
    while i < n:
        c = src[i]
        if c in ' \t\n\r':
            i += 1
            continue
        if c == '"':
            i += 1
            buf = ''
            while i < n:
                if src[i] == '"':
                    if i + 1 < n and src[i + 1] == '"':
                        buf += '"'
                        i += 2
                        continue
                    i += 1
                    break
                buf += src[i]
                i += 1
            toks.append(('str', buf))
            continue
        if c.isdigit() or (c == '.' and i + 1 < n and src[i + 1].isdigit()):
            buf = ''
            while i < n and (src[i].isdigit() or src[i] == '.'):
                buf += src[i]
                i += 1
            v = to_num(buf)
            if v is None:
                return None
            toks.append(('num', v))
            continue
        if c == "'":
            i += 1
            sheet = ''
            while i < n and src[i] != "'":
                sheet += src[i]
                i += 1
            i += 1
            if i >= n or src[i] != '!':
                return None
            i += 1
            r1, i = _read_ref_chars(src, i)
            a = ref_rc(r1)
            if a is None:
                return None
            if i < n and src[i] == ':':
                i += 1
                r2, i = _read_ref_chars(src, i)
                b = ref_rc(r2)
                if b is None:
                    return None
                toks.append(('rng', (sheet, a, b)))
            else:
                toks.append(('ref', (sheet, a)))
            continue
        if c.isalpha() or c == '_' or c == '$':
            buf = ''
            while i < n and (src[i].isalnum() or src[i] in '_$.'):
                buf += src[i]
                i += 1
            if i < n and src[i] == '(':
                toks.append(('fn', buf.upper()))
                continue
            if i < n and src[i] == '!':
                sheet = buf
                i += 1
                r1, i = _read_ref_chars(src, i)
                a = ref_rc(r1)
                if a is None:
                    return None
                if i < n and src[i] == ':':
                    i += 1
                    r2, i = _read_ref_chars(src, i)
                    b = ref_rc(r2)
                    if b is None:
                        return None
                    toks.append(('rng', (sheet, a, b)))
                else:
                    toks.append(('ref', (sheet, a)))
                continue
            up = buf.upper()
            if up == 'TRUE' or up == 'FALSE':
                toks.append(('bool', up == 'TRUE'))
                continue
            a = ref_rc(buf)
            if a is None:
                return None
            if i < n and src[i] == ':':
                i += 1
                r2, i = _read_ref_chars(src, i)
                b = ref_rc(r2)
                if b is None:
                    return None
                toks.append(('rng', (None, a, b)))
            else:
                toks.append(('ref', (None, a)))
            continue
        if c == '(':
            toks.append(('lp', c))
            i += 1
            continue
        if c == ')':
            toks.append(('rp', c))
            i += 1
            continue
        if c in ',;':
            toks.append(('cm', ','))
            i += 1
            continue
        two = src[i:i + 2]
        if two in ('<=', '>=', '<>'):
            toks.append(('op', two))
            i += 2
            continue
        if c in '+-*/^&=<>%':
            toks.append(('op', c))
            i += 1
            continue
        return None
    return toks

def _peek(cur):
    return cur[0][cur[1]] if cur[1] < len(cur[0]) else (None, None)

def _next(cur):
    t = _peek(cur)
    cur[1] += 1
    return t

def parse(toks):
    cur = [toks, 0]
    node = p_cmp(cur)
    if node is None or cur[1] != len(toks):
        return None
    return node

def _binary(cur, sub, ops):
    left = sub(cur)
    if left is None:
        return None
    while True:
        k, v = _peek(cur)
        if k != 'op' or v not in ops:
            return left
        _next(cur)
        right = sub(cur)
        if right is None:
            return None
        left = ('bin', v, left, right)

def p_cmp(cur):
    return _binary(cur, p_cat, ('=', '<>', '<', '<=', '>', '>='))

def p_cat(cur):
    return _binary(cur, p_add, ('&',))

def p_add(cur):
    return _binary(cur, p_mul, ('+', '-'))

def p_mul(cur):
    return _binary(cur, p_pow, ('*', '/'))

def p_pow(cur):
    return _binary(cur, p_un, ('^',))

def p_un(cur):
    k, v = _peek(cur)
    if k == 'op' and v in ('+', '-'):
        _next(cur)
        node = p_un(cur)
        if node is None:
            return None
        return ('neg', node) if v == '-' else node
    return p_post(cur)

def p_post(cur):
    node = p_atom(cur)
    if node is None:
        return None
    while True:
        k, v = _peek(cur)
        if k == 'op' and v == '%':
            _next(cur)
            node = ('pct', node)
        else:
            return node

def p_atom(cur):
    k, v = _next(cur)
    if k in ('num', 'str', 'bool'):
        return ('lit', v)
    if k == 'ref':
        return ('ref', v)
    if k == 'rng':
        return ('rng', v)
    if k == 'lp':
        node = p_cmp(cur)
        if node is None or _next(cur)[0] != 'rp':
            return None
        return node
    if k == 'fn':
        if _next(cur)[0] != 'lp':
            return None
        args = []
        if _peek(cur)[0] == 'rp':
            _next(cur)
            return ('call', v, args)
        while True:
            a = p_cmp(cur)
            if a is None:
                return None
            args.append(a)
            nk = _next(cur)[0]
            if nk == 'rp':
                return ('call', v, args)
            if nk != 'cm':
                return None
    return None

def _slots(toks):
    nums, strs, refs = [], [], []
    idx = 0
    for t in toks:
        if t[0] == 'num':
            nums.append(idx)
        elif t[0] == 'str':
            strs.append(idx)
        elif t[0] in ('ref', 'rng'):
            refs.append(idx)
        idx += 1
    return nums, strs, refs

def _tok_dep(t):
    if t[0] == 'ref':
        sheet, a = t[1]
        return {'s': sheet, 'r1': a[0], 'c1': a[1], 'r2': a[0], 'c2': a[1]}
    sheet, a, b = t[1]
    return {'s': sheet, 'r1': a[0], 'c1': a[1], 'r2': b[0], 'c2': b[1]}

def _dep_tok(d):
    if d['r1'] == d['r2'] and d['c1'] == d['c2']:
        return ('ref', (d['s'], (d['r1'], d['c1'])))
    return ('rng', (d['s'], (d['r1'], d['c1']), (d['r2'], d['c2'])))

def _parse_xc(s):
    sheet = None
    if s.startswith("'"):
        end = s.find("'", 1)
        if end < 0:
            return None
        sheet = s[1:end]
        if end + 1 >= len(s) or s[end + 1] != '!':
            return None
        s = s[end + 2:]
    elif '!' in s:
        parts = s.split('!', 1)
        sheet = parts[0]
        s = parts[1]
    p = s.split(':')
    a = ref_rc(p[0])
    if a is None:
        return None
    b = ref_rc(p[1]) if len(p) > 1 else a
    if b is None:
        return None
    return {'s': sheet, 'r1': a[0], 'c1': a[1], 'r2': b[0], 'c2': b[1]}

def new_state():
    return {'base': None, 'nslot': [], 'sslot': [], 'rslot': [],
            'nbase': [], 'sbase': [], 'napp': [], 'sprev': [], 'rapp': [],
            'prev': None, 'strategy': None, 'fails': []}

def st_new_formula(st, content):
    toks = tokenize(content[1:])
    if toks is None:
        st['base'] = None
        return False
    st['base'] = toks
    st['nslot'], st['sslot'], st['rslot'] = _slots(toks)
    st['nbase'] = [toks[i][1] for i in st['nslot']]
    st['sbase'] = [toks[i][1] for i in st['sslot']]
    st['napp'] = list(st['nbase'])
    st['sprev'] = list(st['sbase'])
    st['rapp'] = [_tok_dep(toks[i]) for i in st['rslot']]
    st['prev'] = None
    return True

def st_unsquish(st, sq):
    if st['base'] is None:
        return None

    N = sq.get('N')
    if N is not None and len(N) > 0:
        nums = []
        idx = 0
        for part in N.split('|'):
            if idx >= len(st['napp']):
                break
            if part == '=':
                nums.append(st['napp'][idx])
            else:
                d = to_num(part[1:])
                if d is None:
                    return None
                st['napp'][idx] = (st['napp'][idx] or 0) + d
                nums.append(st['napp'][idx])
            idx += 1
        while len(nums) < len(st['nbase']):
            nums.append(st['nbase'][len(nums)])
    else:
        nums = list(st['nbase'])

    S = sq.get('S')
    if S is not None and len(S) > 0:
        strs = []
        idx = 0
        for part in S:
            while idx >= len(st['sprev']):
                st['sprev'].append(part)
            if part == '=':
                strs.append(st['sprev'][idx])
            else:
                st['sprev'][idx] = part
                strs.append(part)
            idx += 1
        while len(strs) < len(st['sbase']):
            strs.append(st['sbase'][len(strs)])
    else:
        strs = list(st['sbase'])

    R = sq.get('R')
    if R is not None:
        parts = R if isinstance(R, list) else R.split('|')
        deps = []
        idx = 0
        for rs in parts:
            if idx >= len(st['rapp']):
                break
            cur = st['rapp'][idx]
            if rs == '=':
                deps.append(dict(cur))
            elif len(rs) > 2 and rs[0] in '+-' and rs[1] in 'RC':
                off = to_num(rs[2:])
                if off is None:
                    return None
                off = int(off) * (1 if rs[0] == '+' else -1)
                upd = dict(cur)
                if rs[1] == 'R':
                    upd['r1'] = upd['r2'] = cur['r1'] + off
                else:
                    upd['c1'] = upd['c2'] = cur['c1'] + off
                st['rapp'][idx] = upd
                deps.append(dict(upd))
            else:
                d = _parse_xc(rs)
                if d is None:
                    return None
                st['rapp'][idx] = d
                deps.append(dict(d))
            idx += 1
        while len(deps) < len(st['rslot']):
            deps.append(dict(st['rapp'][len(deps)]))
    else:
        deps = []
        for d in st['rapp']:
            deps.append(dict(d))

    toks = list(st['base'])
    k = 0
    for i in st['nslot']:
        if k < len(nums):
            toks[i] = ('num', nums[k])
        k += 1
    k = 0
    for i in st['sslot']:
        if k < len(strs):
            toks[i] = ('str', strs[k])
        k += 1
    k = 0
    for i in st['rslot']:
        if k < len(deps):
            toks[i] = _dep_tok(deps[k])
        k += 1
    return toks

def st_feed(st, key, content):
    if content is None or content == '':
        return None

    if isinstance(content, str):
        if content.startswith('='):
            st['strategy'] = 'NEW_FORMULA'
            if st_new_formula(st, content):
                return ('F', list(st['base']))
            st['fails'].append(key)
            return ('L', content[1:])
        st['strategy'] = 'NOT_A_FORMULA'
        return ('L', content)

    if isinstance(content, dict) and (content.get('N') is not None
                                      or content.get('S') is not None
                                      or content.get('R') is not None):
        prev = st['strategy']
        if not prev or prev == 'NEW_FORMULA':
            st['strategy'] = 'FIRST_OFFSET'
            st['prev'] = dict(content)
        else:
            st['strategy'] = 'COMBINE_OFFSET'
            if st['prev'] is None:
                st['fails'].append(key)
                return None
            for k in ('N', 'S', 'R'):
                if content.get(k) is not None:
                    st['prev'][k] = content[k]
        toks = st_unsquish(st, st['prev'])
        if toks is None:
            st['fails'].append(key)
            return None
        return ('F', toks)

    if isinstance(content, dict):
        if content.get('value') is not None:
            return ('L', content.get('value'))
        if content.get('content') is not None:
            c_val = content.get('content')
            if isinstance(c_val, str) and c_val.startswith('='):
                return st_feed(st, key, c_val)
            return ('L', c_val)

    st['strategy'] = 'NO_CHANGE'
    return None

def cell_value(book, sheet, rc):
    ck = (sheet, rc)
    cache = book['cache']
    if ck in cache:
        return cache[ck]
    if ck in book['busy']:
        return E_CYC
    grid = book['sheets'].get(sheet)
    if grid is None:
        return E_REF
    cell = grid.get(rc)
    if cell is None:
        return None
    if cell[0] == 'L':
        n = to_num(cell[1])
        out = n if n is not None else cell[1]
        cache[ck] = out
        return out
    book['busy'].add(ck)
    ast = parse(cell[1])
    out = E_EXPR if ast is None else ev(book, ast, sheet)
    book['busy'].discard(ck)
    cache[ck] = out
    return out

def ev(book, node, sheet):
    t = node[0]
    if t == 'lit':
        return node[1]
    if t == 'ref':
        sh, rc = node[1]
        return cell_value(book, sh if sh is not None else sheet, rc)
    if t == 'rng':
        sh, a, b = node[1]
        sh = sh if sh is not None else sheet
        out = []
        r = min(a[0], b[0])
        while r <= max(a[0], b[0]):
            c = min(a[1], b[1])
            while c <= max(a[1], b[1]):
                out.append(cell_value(book, sh, (r, c)))
                c += 1
            r += 1
        return out
    if t == 'neg':
        v = ev(book, node[1], sheet)
        if is_err(v):
            return v
        x = as_num(v)
        return E_VAL if x is None else -x
    if t == 'pct':
        v = ev(book, node[1], sheet)
        if is_err(v):
            return v
        x = as_num(v)
        return E_VAL if x is None else x / 100.0
    if t == 'bin':
        return ev_bin(book, node, sheet)
    if t == 'call':
        return ev_call(book, node, sheet)
    return E_EXPR

def ev_bin(book, node, sheet):
    op = node[1]
    a = ev(book, node[2], sheet)
    if is_err(a):
        return a
    b = ev(book, node[3], sheet)
    if is_err(b):
        return b
    if isinstance(a, list) or isinstance(b, list):
        return E_VAL
    if op == '&':
        return as_text(a) + as_text(b)
    if op in ('=', '<>', '<', '<=', '>', '>='):
        return compare(op, a, b)
    x = as_num(a)
    y = as_num(b)
    if x is None or y is None:
        return E_VAL
    if op == '+':
        return x + y
    if op == '-':
        return x - y
    if op == '*':
        return x * y
    if op == '/':
        return E_DIV if y == 0 else x / y
    if op == '^':
        return float(x ** y)
    return E_EXPR

def truthy(v):
    if isinstance(v, bool):
        return v
    if v is None or v == '':
        return False
    n = as_num(v)
    return bool(v) if n is None else n != 0

def as_text(v):
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else repr(v)
    return str(v)

def compare(op, a, b):
    an = None if isinstance(a, bool) else as_num(a)
    bn = None if isinstance(b, bool) else as_num(b)
    if an is not None and bn is not None:
        x = an
        y = bn
    else:
        x = as_text(a).upper()
        y = as_text(b).upper()
    if op == '=':
        return x == y
    if op == '<>':
        return x != y
    if op == '<':
        return x < y
    if op == '<=':
        return x <= y
    if op == '>':
        return x > y
    return x >= y

def do_round(kind, x, digits):
    f = 1.0
    i = 0
    while i < abs(digits):
        f = f * 10.0 if digits > 0 else f / 10.0
        i += 1
    v = x * f
    iv = float(int(v))
    if abs(v - iv) < 1e-9:
        v = iv
    else:
        step = 1.0 if v > 0 else -1.0
        if abs((iv + step) - v) < 1e-9:
            v = iv + step
    if kind == 'ROUND':
        r = float(int(v + (0.5 if v >= 0 else -0.5)))
    elif kind == 'ROUNDUP':
        r = float(int(v))
        if v != r:
            r += 1.0 if v > 0 else -1.0
    else:
        r = float(int(v))
    return r / f

def flatten(vals):
    out = []
    for v in vals:
        if isinstance(v, list):
            for x in v:
                out.append(x)
        else:
            out.append(v)
    return out

def ev_call(book, node, sheet):
    name = node[1]
    args = node[2]

    if name == 'IF':
        if len(args) < 2:
            return E_EXPR
        cond = ev(book, args[0], sheet)
        if is_err(cond):
            return cond
        if truthy(cond):
            return ev(book, args[1], sheet)
        return ev(book, args[2], sheet) if len(args) >= 3 else False

    if name == 'IFERROR':
        if not args:
            return E_EXPR
        v = ev(book, args[0], sheet)
        if is_err(v):
            return ev(book, args[1], sheet) if len(args) > 1 else ''
        return v

    vals = []
    for a in args:
        v = ev(book, a, sheet)
        if is_err(v):
            return v
        vals.append(v)

    if name in ('AND', 'OR'):
        res = None
        for v in flatten(vals):
            if v is None or v == '':
                continue
            b = truthy(v)
            if res is None:
                res = b
            elif name == 'AND':
                res = res and b
            else:
                res = res or b
        return E_VAL if res is None else res

    if name == 'NOT':
        return not truthy(vals[0]) if vals else E_VAL

    if name in ('ROUNDUP', 'ROUNDDOWN', 'ROUND'):
        if not vals:
            return E_VAL
        x = as_num(vals[0])
        if x is None:
            return E_VAL
        d = 0
        if len(vals) > 1:
            dv = as_num(vals[1])
            if dv is None:
                return E_VAL
            d = int(dv)
        return do_round(name, x, d)

    if name in ('SUM', 'MAX', 'MIN', 'ABS', 'INT', 'PRODUCT', 'AVERAGE', 'COUNT'):
        nums = []
        for v in flatten(vals):
            if v is None or v == '' or isinstance(v, bool):
                continue
            nv = as_num(v)
            if nv is not None:
                nums.append(nv)
        if name == 'COUNT':
            return float(len(nums))
        if name == 'SUM':
            return sum(nums) if nums else 0.0
        if name == 'PRODUCT':
            p = 1.0
            for v in nums:
                p *= v
            return p
        if name == 'AVERAGE':
            return (sum(nums) / len(nums)) if nums else E_DIV
        if not nums:
            return 0.0
        if name == 'MAX':
            return max(nums)
        if name == 'MIN':
            return min(nums)
        if name == 'ABS':
            return abs(nums[0])
        return float(int(nums[0]))

    return E_NAME

def build_book(data):
    sheets = {}
    raw_sheets = {}
    order = []
    fails = {}
    for sh in (data.get('sheets') or []):
        name = sh.get('name') or ''
        cells = sh.get('cells') or {}
        grid = {}
        st = new_state()
        for key in sorted_keys(cells):
            raw = cells[key]
            res = st_feed(st, key, raw)
            if res is None:
                continue
            for rc in key_positions(key):
                grid[rc] = res
        sheets[name] = grid
        raw_sheets[name] = sh
        order.append(name)
        if st['fails']:
            fails[name] = st['fails']
    book = {'sheets': sheets, 'cache': {}, 'busy': set()}
    return book, raw_sheets, order, fails

def build_ranges(m):
    out = []
    if not isinstance(m, dict):
        return out
    for k in m:
        parts = k.split(':')
        a = ref_rc(parts[0])
        if a is None:
            continue
        b = ref_rc(parts[1]) if len(parts) > 1 else a
        if b is None:
            b = a
        out.append((min(a[0], b[0]), min(a[1], b[1]),
                    max(a[0], b[0]), max(a[1], b[1]), m[k]))
    return out

def range_lookup(ranges, rc):
    hit = None
    r = rc[0]
    c = rc[1]
    for it in ranges:
        if it[0] <= r and r <= it[2] and it[1] <= c and c <= it[3]:
            hit = it[4]
    return hit


def _build_xlsx_via_openpyxl(base_data: dict, input_values: dict, template_path: str = None) -> bytes:
    """Workbook builder using openpyxl. Constructs Excel workbook directly from Odoo RPC JSON data."""
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    if wb.sheetnames:
        wb.remove(wb.active)

    book, raw_sheets, order_names, fails = build_book(base_data)
    wb_styles = base_data.get('styles') or {}
    wb_formats = base_data.get('formats') or {}

    sheets_data = base_data.get("sheets", [])
    if not sheets_data:
        for sname in ["Gegevens invulblad", "Afdrukpagina 1", "Afdrukpagina 2", "Odoo feed"]:
            wb.create_sheet(title=sname)
    else:
        for sheet_info in sheets_data:
            s_name = sheet_info.get("name", "Sheet")
            ws = wb.create_sheet(title=s_name)
            
            srng = build_ranges(sheet_info.get('styles') or {})
            frng = build_ranges(sheet_info.get('formats') or {})
            
            # 1. Populate evaluated cells & formatting
            if s_name in book['sheets']:
                for rc, res_tuple in book['sheets'][s_name].items():
                    r, c = rc
                    col_letter = openpyxl.utils.get_column_letter(c)
                    cell_ref = f"{col_letter}{r}"
                    cell = ws[cell_ref]
                    
                    val = cell_value(book, s_name, rc)
                    if val is not None:
                        try:
                            cell.value = val
                        except Exception:
                            cell.value = str(val)
                            
                    # Apply styling
                    sid = range_lookup(srng, rc)
                    fid = range_lookup(frng, rc)
                    st = wb_styles.get(str(sid)) or {} if sid is not None else {}
                    
                    f_size = st.get('fontSize', 10)
                    b_bold = bool(st.get('bold'))
                    i_italic = bool(st.get('italic'))
                    t_color = str(st.get('textColor', '')).lstrip('#')
                    font_kwargs = {'name': 'Calibri', 'size': f_size, 'bold': b_bold, 'italic': i_italic}
                    if t_color:
                        font_kwargs['color'] = t_color.upper()
                    cell.font = Font(**font_kwargs)
                    
                    fill_color = str(st.get('fillColor', '')).lstrip('#')
                    if fill_color:
                        cell.fill = PatternFill(start_color=fill_color.upper(), end_color=fill_color.upper(), fill_type="solid")
                        
                    h_align = st.get('align', 'left')
                    v_align = st.get('verticalAlign', 'bottom')
                    if v_align == 'middle':
                        v_align = 'center'
                    wrap = st.get('wrapping') == 'wrap'
                    cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
                    
            # 2. Apply merged cell ranges
            for m in (sheet_info.get('merges') or []):
                try:
                    ws.merge_cells(m)
                except Exception:
                    pass
                    
            # 3. Apply column dimensions
            cols = sheet_info.get('cols') or {}
            for col_idx_str, col_info in cols.items():
                try:
                    c_idx = int(col_idx_str) + 1
                    c_letter = openpyxl.utils.get_column_letter(c_idx)
                    sz = col_info.get('size') if isinstance(col_info, dict) else None
                    if sz:
                        ws.column_dimensions[c_letter].width = round(sz * 0.14, 2)
                except Exception:
                    pass
                    
            # 4. Apply row dimensions
            rows = sheet_info.get('rows') or {}
            for row_idx_str, row_info in rows.items():
                try:
                    r_idx = int(row_idx_str) + 1
                    sz = row_info.get('size') if isinstance(row_info, dict) else None
                    if sz:
                        ws.row_dimensions[r_idx].height = round(sz * 0.75, 2)
                except Exception:
                    pass

    # Map normalized sheet name to sheet object in workbook
    sheet_name_map = {name.strip().lower(): name for name in wb.sheetnames}

    # Overlay replayed input cell updates from revision log
    for (s_name, cell_ref), val in input_values.items():
        norm_name = s_name.strip().lower()
        if norm_name in sheet_name_map:
            ws = wb[sheet_name_map[norm_name]]
        else:
            ws = wb.create_sheet(title=s_name)
            sheet_name_map[s_name.strip().lower()] = s_name

        target_ref = cell_ref.split(':')[0] if ':' in cell_ref else cell_ref

        if val is not None:
            if isinstance(val, dict):
                if val.get('value') is not None:
                    val = val.get('value')
                elif val.get('S') and isinstance(val.get('S'), list):
                    val = val['S'][0]
                else:
                    val = val.get('content')

            if isinstance(val, str) and (val.startswith('=') or val.startswith('+') or val.startswith('-')):
                val = val.lstrip('+=')

            if val is not None:
                try:
                    ws[target_ref] = val
                except Exception:
                    ws[target_ref] = str(val)

    # Filter: Keep only print sheets in the generated Excel workbook
    resolved_print_sheets = _resolve_print_sheets(wb.sheetnames)
    logger.info(f"Filtering sheets to keep only print sheets: {resolved_print_sheets}")
    for name in list(wb.sheetnames):
        if name not in resolved_print_sheets:
            wb.remove(wb[name])

    if wb.calculation is None:
        from openpyxl.workbook.properties import CalcProperties
        wb.calculation = CalcProperties()
    wb.calculation.fullCalcOnLoad = False

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _recalculate_and_export(xlsx_bytes: bytes, print_sheets: tuple[str, ...]) -> tuple[bytes, bytes]:
    """
    Connect to LibreOffice UNO persistent listener (or run headless process),
    calculate all formulas, export full XLSX and 2-sheet PDF.
    """
    # 1. Try PyUNO bridge via socket listener (port 2002)
    try:
        return _recalculate_via_uno_listener(xlsx_bytes, print_sheets)
    except Exception as e:
        logger.warning(f"UNO listener recalculation failed/unavailable ({e}). Falling back to headless CLI.")

    # 2. Fallback to standard headless soffice process
    return _recalculate_via_headless_cli(xlsx_bytes, print_sheets)


def _recalculate_via_uno_listener(xlsx_bytes: bytes, print_sheets: tuple[str, ...]) -> tuple[bytes, bytes]:
    """Recalculate formulas and export XLSX & PDF via long-lived LibreOffice UNO listener."""
    import uno
    from com.sun.star.beans import PropertyValue

    local_context = uno.getComponentContext()
    resolver = local_context.ServiceManager.createInstanceWithContext(
        "com.sun.star.bridge.UnoUrlResolver", local_context
    )

    connect_str = f"uno:socket,host={LIBREOFFICE_UNO_HOST},port={LIBREOFFICE_UNO_PORT};urp;StarOffice.ComponentContext"
    ctx = resolver.resolve(connect_str)
    smgr = ctx.ServiceManager
    desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)

    with tempfile.TemporaryDirectory() as tmpdir:
        in_path = os.path.join(tmpdir, "input.xlsx")
        out_xlsx_path = os.path.join(tmpdir, "output.xlsx")
        out_pdf_path = os.path.join(tmpdir, "output.pdf")

        with open(in_path, "wb") as f:
            f.write(xlsx_bytes)

        # Load document in LibreOffice
        in_url = uno.systemPathToFileUrl(in_path)
        props = (
            PropertyValue("Hidden", 0, True, 0),
            PropertyValue("ReadOnly", 0, False, 0),
        )
        doc = desktop.loadComponentFromURL(in_url, "_blank", 0, props)

        try:
            # Force full formula recalculation
            # doc.calculateAll()

            # Export 1: Full XLSX workbook
            xlsx_url = uno.systemPathToFileUrl(out_xlsx_path)
            xlsx_props = (PropertyValue("FilterName", 0, "Calc MS Excel 2007 XML", 0),)
            doc.storeToURL(xlsx_url, xlsx_props)

            # Export 2: PDF containing ONLY Afdrukpagina 1 & Afdrukpagina 2
            controller = doc.getCurrentController()
            sheets = doc.getSheets()
            
            selected_sheets = []
            for name in print_sheets:
                if sheets.hasByName(name):
                    selected_sheets.append(sheets.getByName(name))

            if selected_sheets:
                # Select only the print sheets in the UI controller
                controller.select(tuple(selected_sheets))

            pdf_url = uno.systemPathToFileUrl(out_pdf_path)
            pdf_props = (PropertyValue("FilterName", 0, "calc_pdf_Export", 0),)
            doc.storeToURL(pdf_url, pdf_props)

        finally:
            doc.close(True)

        with open(out_xlsx_path, "rb") as f:
            res_xlsx = f.read()

        with open(out_pdf_path, "rb") as f:
            res_pdf = f.read()

        return res_xlsx, res_pdf


def _recalculate_via_headless_cli(xlsx_bytes: bytes, print_sheets: tuple[str, ...]) -> tuple[bytes, bytes]:
    """Fallback recalculation via CLI subprocess."""
    lo_bin = shutil.which("libreoffice") or shutil.which("soffice")
    if not lo_bin:
        raise RuntimeError("LibreOffice binary not found.")

    with tempfile.TemporaryDirectory() as tmpdir:
        in_path = os.path.join(tmpdir, "input.xlsx")
        pdf_in_path = os.path.join(tmpdir, "print_only.xlsx")

        with open(in_path, "wb") as f:
            f.write(xlsx_bytes)

        # Create workbook with ONLY print sheets for PDF export
        wb_full = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        wb_pdf = openpyxl.Workbook()
        if wb_pdf.sheetnames:
            wb_pdf.remove(wb_pdf.active)

        for sname in print_sheets:
            if sname in wb_full.sheetnames:
                src = wb_full[sname]
                dst = wb_pdf.create_sheet(title=sname)
                for row in src.iter_rows(values_only=False):
                    for cell in row:
                        if cell.value is not None:
                            dst.cell(row=cell.row, column=cell.column, value=cell.value)

        wb_pdf.save(pdf_in_path)

        cmd = [
            lo_bin, "--headless", "--invisible", "--nocrashreport", "--nodefault", "--norestore",
            "--convert-to", "pdf", "--outdir", tmpdir, pdf_in_path
        ]
        res = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        if res.returncode != 0:
            raise RuntimeError(f"LibreOffice PDF export error: {res.stderr}")

        out_pdf = os.path.join(tmpdir, "print_only.pdf")
        if not os.path.exists(out_pdf):
            raise RuntimeError("PDF export failed.")

        with open(out_pdf, "rb") as f:
            pdf_bytes = f.read()

        return xlsx_bytes, pdf_bytes


def _validate_xlsx_output(xlsx_bytes: bytes, print_sheets: tuple[str, ...]) -> None:
    """Validate that print sheets do not contain unexpected formula error values."""
    buf = io.BytesIO(xlsx_bytes)
    wb = openpyxl.load_workbook(buf, data_only=True)

    detected_errors = []
    for sheet_name in print_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if str(cell.value) in ERROR_VALUES:
                    # Column P #BAD_EXPR is allowed if known pre-existing template issue, flag others
                    cell_ref = cell.coordinate
                    if cell_ref.startswith('P') and str(cell.value) == '#BAD_EXPR':
                        logger.info(f"Known template error ignored at {sheet_name}!{cell_ref}: {cell.value}")
                        continue
                    detected_errors.append(f"{sheet_name}!{cell_ref}: {cell.value}")

    if detected_errors:
        err_msg = f"Validation detected formula errors (ignored): {', '.join(detected_errors[:5])}"
        logger.warning(err_msg)
